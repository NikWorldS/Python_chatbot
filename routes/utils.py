from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import mimetypes
import calendar
import datetime
import openpyxl
import smtplib
import shutil
import math
import time
import os
from dotenv import load_dotenv, find_dotenv

load_dotenv(find_dotenv())
from_mail = os.getenv("FROM_EMAIL")
password = os.getenv("EMAIL_PASS")


def generate():
    next_answer = math.floor(time.time()) + (60 * 60 * 24)
    return next_answer


def create_list_payers(class_name):
    x = 13
    payers_list = []
    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb[wb.sheetnames[0]]

    while sheet[f'b{x}'].value.lower() != 'итого':
        if sheet[f"b{x}"].value == sheet[f"b{x + 1}"].value:
            payers_list.append(((sheet[f'b{x}']).value + (sheet[f'c{x}']).value).strip())
            print(1)
        else:
            payers_list.append(((sheet[f'b{x}']).value).strip())
        x += 1
    return payers_list


def filling_template(values_dict, class_name):
    x = 13
    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb[wb.sheetnames[0]]
    week_date = datetime.datetime.today().weekday()
    columns = ['f', 'g', 'h', 'i', 'j', 'k']
    sheet[f'{columns[week_date]}12'] = datetime.datetime.now().strftime('%d')
    for value in values_dict:
        sheet.cell(row=x, column=(6 + week_date)).value = value
        x += 1
        wb.save(f'attachments/Template_{class_name}.xlsx')
        wb.close()


def create_template(doc):
    x = 13
    columns = ['f', 'g', 'h', 'i', 'j', 'k']  # Столбцы, в которых записываются данные об отсутствующих

    try:

        wb = openpyxl.load_workbook(filename=f'{doc}')
        sheet = wb[wb.sheetnames[0]]

        sheet['F7'] = None
        sheet['H7'] = None
        sheet['K6'] = None
        sheet['L9'] = None
        sheet['K9'] = '=F12'

        while sheet[f'b{x}'].value != 'итого':
            x += 1
        sheet_offset = sheet['b13'].offset(row=x - 13, column=0)

        for i in columns:  # Запись формул по сложению в строку
            formula = f'=SUM({i}{13}:{i}{sheet_offset.row - 1})'
            sheet[f'{i}{sheet_offset.row}'] = formula

        for i in range(sheet_offset.row - 1, 12, -1):  # Запись формул по сложению в столбец
            formula2 = f'=SUM(f{i}:k{i})'
            sheet[f'l{i}'] = formula2

        formula3 = '=IF('  # Запись формулы по проверке и вычислении итого
        for i in columns:
            formula3 += f'+{i}{sheet_offset.row}'
        formula3 += '='
        for j in range(sheet_offset.row - 1, 12, -1):
            formula3 += f'+l{j}'
        formula3 += ','
        for i in columns:
            formula3 += f'+{i}{sheet_offset.row}'
        formula3 += ',"ошибка")'
        sheet[f'l{sheet_offset.row}'] = formula3

        for i in columns:  # очистка ячеек
            for j in range(sheet_offset.row - 1, 11, -1):
                sheet[f'{i}{j}'] = None

        class_name = sheet['d3'].value.strip()
        class_name = "".join(class_name.split())
        sheet['d3'] = class_name
        wb.save(f'attachments/Template_{class_name}.xlsx')
        wb.close()
        return "Файл принят"

    except Exception as create_error:
        return f'Во время создания табеля возникла ошибка ({create_error})'
    finally:
        os.remove(doc)


def clear_template(class_name, date_today):
    x = 13

    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb[wb.sheetnames[0]]

    columns = ['f', 'g', 'h', 'i', 'j', 'k']

    while sheet[f'b{x}'].value != 'итого':
        x += 1
    sheet_offset = sheet['b13'].offset(row=x - 13, column=0)

    for i in columns:  # очистка ячеек
        for j in range(sheet_offset.row - 1, 11, -1):
            sheet[f'{i}{j}'] = None

    wb.save(f'attachments/Template_{class_name}.xlsx')
    wb.close()

    if not os.path.isdir(f"archive/class_{class_name}"):
        os.mkdir(f"archive/class_{class_name}")

    shutil.copy(f'attachments/Tabel_{class_name}_{date_today}.xlsx',
                f'archive/class_{class_name}/Tabel_{class_name}_{date_today}.xlsx')
    os.remove(f'attachments/Tabel_{class_name}_{date_today}.xlsx')


def send_tabel(class_name=None, mail=None):
    date_today = datetime.datetime.today().date().strftime('%d.%m.%y')
    day = datetime.datetime.today().date().strftime('%d')

    msg = MIMEMultipart()

    message = f'Табель на {date_today}'
    msg["Subject"] = f'Табель {class_name} по питанию на {date_today}'
    to_email = mail
    msg.attach(MIMEText(message, 'plain'))

    server = smtplib.SMTP('smtp.gmail.com: 587')
    server.starttls()
    try:

        wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
        sheet = wb[wb.sheetnames[0]]
        sheet['L9'] = day
        month_num = datetime.datetime.now().month
        month_name = calendar.month_name[month_num]
        month_name_russian = {
            'January': 'Январь',
            'February': 'Февраль',
            'March': 'Март',
            'April': 'Апрель',
            'May': 'Май',
            'June': 'Июнь',
            'July': 'Июль',
            'August': 'Август',
            'September': 'Сентябрь',
            'October': 'Октябрь',
            'November': 'Ноябрь',
            'December': 'Декабрь'
        }[month_name]
        sheet['K6'] = month_name_russian
        sheet["H7"] = date_today

        wb.save(f'attachments/Tabel_{class_name}_{date_today}.xlsx')
        wb.close()

        server.login(from_mail, password)

        file = f'Tabel_{class_name}_{date_today}.xlsx'
        filename = os.path.basename(file)
        ftype, encoding = mimetypes.guess_type(file)
        file_type, subtype = ftype.split('/')

        if file_type == "application":
            with open(f'./attachments/{file}', 'rb') as f:
                file = MIMEApplication(f.read(), subtype)

            file.add_header('content-disposition', 'attachment', filename=filename)
            msg.attach(file)

        clear_template(class_name, date_today)
        server.sendmail(from_mail, to_email, msg.as_string())

        server.close()
        return f'Табель отправлен на почту'

    except Exception as _ex_error:
        return f'Во время отправки произошла ошибка: {_ex_error}'